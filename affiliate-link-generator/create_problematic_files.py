#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
問題のあるファイル形式のテストファイル作成
"""

import os
import xlwt  # 古いxls形式作成用
from openpyxl import Workbook

def create_xls_file():
    """
    古い.xls形式のファイルを作成
    """
    try:
        import xlwt
        print("📄 test.xls を作成中...")
        
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet('商品リスト')
        
        # ヘッダー
        worksheet.write(0, 0, '商品番号')
        worksheet.write(0, 1, '商品名')  
        worksheet.write(0, 2, 'Amazonリンク')
        worksheet.write(0, 3, '楽天リンク')
        
        # データ
        products = ['古いExcel形式テスト', 'xls形式商品', '互換性テスト商品']
        for i, product in enumerate(products):
            worksheet.write(i+1, 0, i+1)
            worksheet.write(i+1, 1, product)
        
        workbook.save('test.xls')
        print("✅ test.xls 作成完了")
        return True
        
    except ImportError:
        print("⚠️ xlwt パッケージが見つかりません")
        print("   pip install xlwt でインストールしてください")
        return False
    except Exception as e:
        print(f"❌ test.xls 作成エラー: {e}")
        return False

def create_corrupted_file():
    """
    破損したファイルを模擬
    """
    print("📄 corrupted.xlsx を作成中...")
    
    # 不正なバイナリデータを書き込み
    with open('corrupted.xlsx', 'wb') as f:
        f.write(b'PK\x03\x04' + b'\x00' * 100)  # 不完全なzipヘッダー
    
    print("✅ corrupted.xlsx 作成完了")

def create_password_protected_file():
    """
    パスワード保護されたファイルを模擬作成
    （実際のパスワード保護は複雑なので、ファイル名だけ作成）
    """
    print("📄 password_protected.xlsx を作成中...")
    
    # 通常のExcelファイルを作成してパスワード保護と見せかける
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "このファイルはパスワード保護されています"
    ws['B1'] = "テスト用商品"
    wb.save('password_protected.xlsx')
    
    print("✅ password_protected.xlsx 作成完了")

def create_wps_mock_file():
    """
    WPS Spreadsheetsファイルを模擬
    """
    print("📄 test.et を作成中...")
    
    # .etファイル（WPS形式）は特殊なので、通常のxlsxを.etにリネーム
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "商品番号"
    ws['B1'] = "商品名"
    ws['B2'] = "WPS形式テスト商品"
    ws['B3'] = "ET形式商品"
    
    # 一旦xlsxとして保存してからリネーム
    wb.save('temp_wps.xlsx')
    
    if os.path.exists('temp_wps.xlsx'):
        os.rename('temp_wps.xlsx', 'test.et')
        print("✅ test.et 作成完了")
    else:
        print("❌ test.et 作成失敗")

def create_large_file():
    """
    大きなファイルを作成（50MB制限テスト用）
    """
    print("📄 large.xlsx を作成中...")
    
    wb = Workbook()
    ws = wb.active
    
    # ヘッダー
    ws['A1'] = "商品番号"
    ws['B1'] = "商品名"
    
    # 大量のデータを追加
    for i in range(2, 10000):  # 約1万行
        ws[f'A{i}'] = i-1
        ws[f'B{i}'] = f"大量データテスト商品{i-1}"
        if i % 1000 == 0:
            print(f"   {i}行目まで作成中...")
    
    wb.save('large.xlsx')
    size = os.path.getsize('large.xlsx') / 1024 / 1024
    print(f"✅ large.xlsx 作成完了 ({size:.1f} MB)")

def main():
    """
    メイン関数
    """
    print("🔧 問題のあるファイル形式のテスト用ファイルを作成中...")
    print("=" * 60)
    
    # 各種問題ファイルを作成
    create_xls_file()
    create_wps_mock_file()
    create_corrupted_file()
    create_password_protected_file()
    create_large_file()
    
    print("\n📋 作成されたファイル一覧:")
    test_files = [
        'test.xls', 'test.et', 'corrupted.xlsx', 
        'password_protected.xlsx', 'large.xlsx'
    ]
    
    for file in test_files:
        if os.path.exists(file):
            size = os.path.getsize(file)
            if size > 1024 * 1024:
                size_str = f"{size / 1024 / 1024:.1f} MB"
            else:
                size_str = f"{size / 1024:.1f} KB"
            print(f"  ✅ {file} ({size_str})")
        else:
            print(f"  ❌ {file} (作成失敗)")

if __name__ == "__main__":
    main()