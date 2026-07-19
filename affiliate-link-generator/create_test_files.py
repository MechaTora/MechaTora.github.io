#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
テスト用Excelファイル作成スクリプト
"""

from openpyxl import Workbook
import os

def create_test_files():
    """
    様々な形式のテストファイルを作成
    """
    
    # 1. 通常のxlsxファイル
    print("📄 test.xlsx を作成中...")
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "商品リスト"
    
    # ヘッダー
    ws1['A1'] = "商品番号"
    ws1['B1'] = "商品名"
    ws1['C1'] = "Amazonリンク"
    ws1['D1'] = "楽天リンク"
    
    # テストデータ
    test_products = [
        "iPhone 15",
        "MacBook Pro",
        "AirPods Pro",
        "iPad Air"
    ]
    
    for i, product in enumerate(test_products, 2):
        ws1[f'A{i}'] = i-1
        ws1[f'B{i}'] = product
    
    wb1.save("test.xlsx")
    print("✅ test.xlsx 作成完了")
    
    # 2. 問題のあるファイル（a.xlsx）を作成
    print("📄 a.xlsx を作成中...")
    wb2 = Workbook()
    ws2 = wb2.active
    
    # 特殊文字を含むデータで問題を再現
    ws2['A1'] = "番号"
    ws2['B1'] = "商品名"
    ws2['B2'] = "テスト商品①"
    ws2['B3'] = "問題のある商品名　※特殊文字"
    ws2['B4'] = ""  # 空のセル
    ws2['B5'] = None  # Noneセル
    
    wb2.save("a.xlsx")
    print("✅ a.xlsx 作成完了")
    
    # 3. 空のファイル
    print("📄 empty.xlsx を作成中...")
    wb3 = Workbook()
    ws3 = wb3.active
    wb3.save("empty.xlsx")
    print("✅ empty.xlsx 作成完了")
    
    # 4. データが不正な形式のファイル
    print("📄 broken.xlsx を作成中...")
    wb4 = Workbook()
    ws4 = wb4.active
    ws4['A1'] = "これは"
    ws4['B1'] = "不正な"
    ws4['C1'] = "データです"
    # B列以外にデータを入れて問題を作る
    ws4['A2'] = "商品データがB列にない"
    wb4.save("broken.xlsx")
    print("✅ broken.xlsx 作成完了")

if __name__ == "__main__":
    create_test_files()
    print("\n🎉 すべてのテストファイルが作成されました")
    
    print("\n📋 作成されたファイル:")
    for file in ["test.xlsx", "a.xlsx", "empty.xlsx", "broken.xlsx"]:
        if os.path.exists(file):
            size = os.path.getsize(file)
            print(f"  ✅ {file} ({size / 1024:.1f} KB)")
        else:
            print(f"  ❌ {file} (作成失敗)")