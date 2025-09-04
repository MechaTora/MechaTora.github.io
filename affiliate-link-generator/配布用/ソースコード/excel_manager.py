#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel管理モジュール
Excelファイルの読み書きを処理
"""

import os
import shutil
import datetime
from typing import Dict, List, Tuple, Optional, Any
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelManager:
    def __init__(self):
        self.workbook: Optional[Workbook] = None
        self.worksheet = None
        self.file_path = ""
        self.products_data: List[Dict[str, Any]] = []
        
    def load_file(self, file_path: str) -> bool:
        """
        Excelファイルを読み込む
        
        Args:
            file_path (str): 読み込むExcelファイルのパス
            
        Returns:
            bool: 読み込み成功時True、失敗時False
        """
        try:
            # ファイル存在チェック
            if not os.path.exists(file_path):
                print(f"❌ ファイルが見つかりません: {file_path}")
                return False
                
            # ファイル拡張子チェック（WPS対応）
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in ['.xlsx', '.xls', '.et', '.ett']:
                print(f"❌ サポートされていないファイル形式: {file_ext}")
                if file_ext in ['.et', '.ett']:
                    print("   → WPS Spreadsheets形式(.et)が検出されました")
                    print("   → .xlsx形式で保存し直してください")
                else:
                    print(f"   → 対応形式: .xlsx, .xls")
                return False
            elif file_ext in ['.et', '.ett']:
                print(f"⚠️ WPS Spreadsheets形式(.et)を検出")
                print("   → 互換モードで読み込みを試行します")
            
            # ファイルアクセス権チェック
            if not os.access(file_path, os.R_OK):
                print(f"❌ ファイル読み取り権限がありません: {file_path}")
                return False
                
            # ファイルサイズチェック（50MB以上は警告）
            file_size = os.path.getsize(file_path)
            if file_size > 50 * 1024 * 1024:  # 50MB
                print(f"⚠️ 大きなファイルです ({file_size / 1024 / 1024:.1f}MB): 処理に時間がかかる場合があります")
            
            # ファイルが使用中でないかチェック
            try:
                with open(file_path, 'r+b'):
                    pass
            except (PermissionError, OSError) as e:
                print(f"❌ ファイルが他のアプリケーションで開かれています: {file_path}")
                print(f"   詳細エラー: {str(e)}")
                return False
            
            print(f"📖 Excelファイルを読み込み中: {os.path.basename(file_path)}")
            print(f"   ファイルサイズ: {file_size / 1024:.1f} KB")
            
            # Excelファイルの読み込み（WPS Spreadsheets対応）
            try:
                # 通常の読み込みを試行
                print("🔄 標準モードで読み込み試行中...")
                self.workbook = load_workbook(file_path, read_only=False, data_only=True)
                print("✅ 標準モードで読み込み成功")
            except Exception as e:
                print(f"⚠️ 標準モードでの読み込みに失敗")
                print(f"   エラー詳細: {str(e)}")
                print(f"   エラー型: {type(e).__name__}")
                
                # WPS Spreadsheets等の互換性問題対応
                try:
                    print("🔄 互換モード1で読み込み試行中...")
                    # data_only=False で再試行
                    self.workbook = load_workbook(file_path, read_only=False, data_only=False)
                    print("✅ 互換モード1で読み込み成功")
                except Exception as e2:
                    print(f"⚠️ 互換モード1での読み込みに失敗")
                    print(f"   エラー詳細: {str(e2)}")
                    try:
                        print("🔄 互換モード2で読み込み試行中...")
                        # read_only=True で再試行
                        self.workbook = load_workbook(file_path, read_only=True, data_only=True)
                        print("✅ 互換モード2で読み込み成功")
                    except Exception as e3:
                        print(f"❌ 全ての読み込み方法が失敗しました")
                        print(f"   最初のエラー: {str(e)}")
                        print(f"   2番目のエラー: {str(e2)}")
                        print(f"   最終エラー: {str(e3)}")
                        print(f"   最終エラー型: {type(e3).__name__}")
                        
                        # 詳細なエラー分析
                        error_msg = str(e3).lower()
                        if "corrupted" in error_msg or "damaged" in error_msg:
                            print("   → ファイルが破損している可能性があります")
                        elif "password" in error_msg:
                            print("   → パスワード保護されたファイルの可能性があります")
                        elif "version" in error_msg or "format" in error_msg:
                            print("   → WPS Spreadsheets形式の可能性があります")
                            print("   → 解決方法: WPS で .xlsx 形式で保存し直してください")
                        elif "wps" in error_msg:
                            print("   → WPS Spreadsheets専用形式です")
                            print("   → 解決方法: Microsoft Excel形式(.xlsx)で保存してください")
                        elif "zipfile" in error_msg or "zip" in error_msg:
                            print("   → ファイル形式が正しくない可能性があります")
                            print("   → Excelで開いて .xlsx 形式で保存し直してください")
                        elif "xml" in error_msg:
                            print("   → XMLファイル構造に問題があります")
                            print("   → Excelで開いて修復後、再保存してください")
                        
                        return False
            
            # ワークシートの取得
            try:
                if self.workbook.worksheets:
                    self.worksheet = self.workbook.active
                    if self.worksheet is None:
                        self.worksheet = self.workbook.worksheets[0]
                else:
                    print("❌ ワークシートが見つかりません")
                    return False
                    
            except Exception as e:
                print(f"❌ ワークシートの取得に失敗: {str(e)}")
                return False
            
            self.file_path = file_path
            
            # データの読み込み
            success = self._load_products_data()
            if success:
                print(f"✅ Excelファイルの読み込み完了: {len(self.products_data)}個の商品データ")
            else:
                print("❌ 商品データの読み込みに失敗")
                
            return success
            
        except Exception as e:
            print(f"❌ 予期しないエラーが発生しました: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _load_products_data(self) -> bool:
        """
        商品データを読み込む
        
        Returns:
            bool: 読み込み成功時True
        """
        try:
            self.products_data = []
            
            # ヘッダー行の検出
            header_row = 1
            max_check_rows = 10  # 最初の10行でヘッダーを探す
            
            for row in range(1, min(max_check_rows + 1, self.worksheet.max_row + 1)):
                cell_b = self.worksheet.cell(row=row, column=2)
                if cell_b.value and isinstance(cell_b.value, str):
                    if any(keyword in str(cell_b.value).lower() for keyword in ['商品', 'product', 'item', 'name']):
                        header_row = row
                        break
                        
            print(f"🔍 ヘッダー行を{header_row}行目として認識")
            
            # データ行の読み込み
            data_start_row = header_row + 1
            product_count = 0
            empty_row_count = 0
            max_empty_rows = 5  # 連続5行空行で終了
            
            for row in range(data_start_row, self.worksheet.max_row + 1):
                try:
                    # B列（商品名）の取得
                    product_name_cell = self.worksheet.cell(row=row, column=2)
                    product_name = product_name_cell.value
                    
                    # 空行チェック
                    if not product_name or str(product_name).strip() == "":
                        empty_row_count += 1
                        if empty_row_count >= max_empty_rows:
                            print(f"🔍 連続{max_empty_rows}行の空行を検出、読み込み終了")
                            break
                        continue
                    else:
                        empty_row_count = 0
                    
                    # 商品名の正規化
                    product_name = str(product_name).strip()
                    
                    # 既存のリンク情報取得
                    amazon_link_cell = self.worksheet.cell(row=row, column=3)
                    rakuten_link_cell = self.worksheet.cell(row=row, column=4)
                    
                    amazon_link = amazon_link_cell.value if amazon_link_cell.value else ""
                    rakuten_link = rakuten_link_cell.value if rakuten_link_cell.value else ""
                    
                    # 商品データの作成
                    product_data = {
                        'row': row,
                        'product_name': product_name,
                        'amazon_link': str(amazon_link).strip(),
                        'rakuten_link': str(rakuten_link).strip(),
                        'processed': False
                    }
                    
                    self.products_data.append(product_data)
                    product_count += 1
                    
                except Exception as e:
                    print(f"⚠️ {row}行目の読み込みでエラー: {str(e)}")
                    continue
            
            if product_count == 0:
                print("❌ 商品データが見つかりません")
                print("   → B列に商品名が入力されているか確認してください")
                return False
                
            print(f"📊 {product_count}個の商品データを読み込みました")
            return True
            
        except Exception as e:
            print(f"❌ 商品データ読み込み中にエラー: {str(e)}")
            return False
    
    def get_products(self) -> List[Dict[str, Any]]:
        """
        商品データのリストを取得
        
        Returns:
            List[Dict]: 商品データのリスト
        """
        return self.products_data.copy()
    
    def update_product_link(self, row: int, column: str, link: str) -> bool:
        """
        商品のリンクを更新
        
        Args:
            row (int): 行番号
            column (str): 列名（'amazon' or 'rakuten'）
            link (str): リンクURL
            
        Returns:
            bool: 更新成功時True
        """
        try:
            if not self.worksheet:
                return False
                
            # 列番号の決定
            col_num = 3 if column == 'amazon' else 4
            
            # セルの更新
            cell = self.worksheet.cell(row=row, column=col_num)
            cell.value = link
            
            # スタイル適用
            if link and link.strip():
                # リンクが設定された場合のスタイル
                cell.font = Font(color="0000FF", underline="single")
                cell.alignment = Alignment(wrap_text=True)
            else:
                # リンクが空の場合は通常スタイル
                cell.font = Font()
                cell.alignment = Alignment()
            
            # データも更新
            for product in self.products_data:
                if product['row'] == row:
                    if column == 'amazon':
                        product['amazon_link'] = link
                    else:
                        product['rakuten_link'] = link
                    break
                    
            return True
            
        except Exception as e:
            print(f"❌ リンク更新エラー (行{row}, {column}): {str(e)}")
            return False
    
    def create_backup(self) -> bool:
        """
        バックアップファイルを作成
        
        Returns:
            bool: バックアップ作成成功時True
        """
        try:
            if not self.file_path or not os.path.exists(self.file_path):
                return False
            
            # バックアップファイル名の生成
            base_name = os.path.splitext(self.file_path)[0]
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{base_name}_backup_{timestamp}.xlsx"
            
            # バックアップ作成
            shutil.copy2(self.file_path, backup_path)
            print(f"💾 バックアップ作成完了: {os.path.basename(backup_path)}")
            return True
            
        except Exception as e:
            print(f"❌ バックアップ作成エラー: {str(e)}")
            return False
    
    def save_file(self, create_backup: bool = True) -> bool:
        """
        ファイルを保存
        
        Args:
            create_backup (bool): バックアップ作成フラグ
            
        Returns:
            bool: 保存成功時True
        """
        try:
            if not self.workbook or not self.file_path:
                print("❌ 保存対象のファイルがありません")
                return False
            
            # バックアップ作成
            if create_backup:
                backup_success = self.create_backup()
                if not backup_success:
                    print("⚠️ バックアップ作成に失敗しましたが、保存を続行します")
            
            # ヘッダー行のスタイル設定
            try:
                for col in range(1, 5):  # A-D列
                    cell = self.worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            except Exception as e:
                print(f"⚠️ ヘッダースタイル設定エラー: {str(e)}")
            
            # 列幅の自動調整
            try:
                column_widths = {
                    1: 8,   # A列: 商品番号
                    2: 30,  # B列: 商品名  
                    3: 25,  # C列: Amazonリンク
                    4: 25   # D列: 楽天リンク
                }
                
                for col, width in column_widths.items():
                    col_letter = get_column_letter(col)
                    self.worksheet.column_dimensions[col_letter].width = width
                    
            except Exception as e:
                print(f"⚠️ 列幅調整エラー: {str(e)}")
            
            # ファイル保存
            print(f"💾 ファイルを保存中: {os.path.basename(self.file_path)}")
            self.workbook.save(self.file_path)
            
            print("✅ ファイル保存完了")
            return True
            
        except Exception as e:
            print(f"❌ ファイル保存エラー: {str(e)}")
            
            # 保存失敗時の対処
            if "permission" in str(e).lower():
                print("   → ファイルが他のアプリケーションで開かれている可能性があります")
            elif "readonly" in str(e).lower():
                print("   → 読み取り専用ファイルの可能性があります")
            
            return False
    
    def get_statistics(self) -> Dict[str, Any]:
        """
        ファイル統計情報を取得
        
        Returns:
            Dict: 統計情報
        """
        if not self.products_data:
            return {}
        
        total_products = len(self.products_data)
        amazon_links = sum(1 for p in self.products_data if p['amazon_link'].strip())
        rakuten_links = sum(1 for p in self.products_data if p['rakuten_link'].strip())
        
        return {
            'total_products': total_products,
            'amazon_links': amazon_links,
            'rakuten_links': rakuten_links,
            'empty_amazon': total_products - amazon_links,
            'empty_rakuten': total_products - rakuten_links
        }
    
    def create_sample_file(self, file_path: str) -> bool:
        """
        サンプルExcelファイルを作成
        
        Args:
            file_path (str): 作成するファイルのパス
            
        Returns:
            bool: 作成成功時True
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "商品リスト"
            
            # ヘッダー設定
            headers = ['商品番号', '商品名', 'Amazonリンク', '楽天リンク']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # サンプルデータ
            sample_products = [
                "iPhone 15 Pro 128GB",
                "ワイヤレスマウス Bluetooth",
                "Python プログラミング本",
                "コーヒーメーカー",
                "Bluetoothスピーカー"
            ]
            
            for row, product in enumerate(sample_products, 2):
                ws.cell(row=row, column=1, value=row-1)  # 商品番号
                ws.cell(row=row, column=2, value=product)  # 商品名
            
            # 列幅調整
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            
            # 保存
            wb.save(file_path)
            print(f"📄 サンプルファイル作成完了: {os.path.basename(file_path)}")
            return True
            
        except Exception as e:
            print(f"❌ サンプルファイル作成エラー: {str(e)}")
            return False

    def close(self):
        """リソースのクリーンアップ"""
        try:
            if self.workbook:
                self.workbook.close()
                self.workbook = None
            self.worksheet = None
            self.file_path = ""
            self.products_data = []
        except Exception as e:
            print(f"⚠️ クリーンアップエラー: {str(e)}")